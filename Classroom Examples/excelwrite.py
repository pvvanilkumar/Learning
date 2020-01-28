from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Benz"

ws.cell(row=1,column=1,value="Hello")
ws.cell(row=1,column=2,value="Bangalore")

for row in range(2,10):
    for column in range(2,10):
        ws.cell(row=row,column=column, value=row*column)

ws.append(["welcome","to","MBRDI"])
ws.append([45,89,34,78])

wb.save("test.xlsx")
