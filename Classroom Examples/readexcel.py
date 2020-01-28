from openpyxl import load_workbook

wb = load_workbook("marks.xlsx")

print(wb.sheetnames)
ws = wb["Marks"]

print(ws.max_row,ws.max_column)

#read a cell
print(ws["A1"].value)
#read a row
row = ws[1]
for cell in row:
    print(cell.value)
#read a column
coll = ws['A']
for cell in coll:
    print(cell.value)
    
for row in ws:
    for cell in row:
        print(cell.value,end = " ")
    print()